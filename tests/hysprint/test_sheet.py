import io

from openpyxl import load_workbook

from sand.hysprint.archive import build_samples
from sand.hysprint.sheet import grid_to_xlsx_bytes, to_sheet

INFO = {
    'project_name': 'perov',
    'batch': 'B1',
    'subbatch': 'a',
    'first_sample': '1',
    'n_samples': 2,
    'date': '2026-08-29',
}


def _archive(steps):
    return {'samples': build_samples(INFO), 'steps': steps}


def test_sheet_lays_out_headers_sections_and_sample_rows():
    archive = _archive(
        [
            {
                'step_type': 'Cleaning UV-Ozone',
                'position_in_experimental_plan': 1,
                'samples': 'all',
                'cleaning_steps': [{'solvent': 'Hellmanex', 'time': 30}],
            },
            {
                'step_type': 'Spin Coating',
                'position_in_experimental_plan': 2,
                'samples': ['perov_B1_a_C-1'],
                'material_name': 'NiOx',
                'rotation_steps': [{'speed': 3000, 'time': 30}],
            },
        ]
    )

    grid, issues = to_sheet(archive)

    assert issues == []
    row0, row1, *rows = grid
    assert row0[0] == 'Experiment Info'
    assert '1: Cleaning UV-Ozone' in row0
    assert '2: Spin Coating' in row0
    assert len(rows) == INFO['n_samples']  # one row per sample
    # every row is rectangular
    assert {len(r) for r in grid} == {len(row0)}
    # the date lands as the ISO string the batch parser accepts
    date_col = row1.index('Date')
    assert rows[0][date_col] == '2026-08-29'
    # the hardcoded substrate defaults fill the substrate block on every
    # row - present-but-empty substrate columns crash the batch parser
    substrate_col = row1.index('Substrate material')
    assert {r[substrate_col] for r in rows} == {'Glass'}
    # the spin coating cells are filled for sample 1 only
    material_col = row1.index('Material name')
    assert rows[0][material_col] == 'NiOx'
    assert rows[1][material_col] == ''


def test_sheet_reports_unrepresentable_fields_as_issues():
    archive = _archive(
        [
            {
                'step_type': 'Cleaning UV-Ozone',
                'position_in_experimental_plan': 1,
                'samples': 'all',
                'made_up_field': 'x',
            }
        ]
    )

    grid, issues = to_sheet(archive)

    assert any('made_up_field' in issue for issue in issues)


def test_sheet_warns_for_unknown_step_type():
    archive = _archive(
        [
            {
                'step_type': 'Underwater Welding',
                'position_in_experimental_plan': 1,
                'samples': 'all',
                'time': 5,
            }
        ]
    )

    grid, issues = to_sheet(archive)

    assert any('no template' in issue for issue in issues)
    assert grid[0][0] == 'Experiment Info'  # only the info section remains


def test_xlsx_bytes_round_trip_with_typed_cells():
    grid = [
        ['Experiment Info', ''],
        ['Batch', 'Number of pixels'],
        ['B1', '6'],
    ]

    data = grid_to_xlsx_bytes(grid)

    ws = load_workbook(io.BytesIO(data)).active
    assert ws['A2'].value == 'Batch'
    assert ws['A3'].value == 'B1'
    pixel_count = 6
    assert ws['B3'].value == pixel_count  # numeric cell, not the string '6'
